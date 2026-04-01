"""
Creates data/test_data.xlsx with sample rows for the Login sheet.
Run once to bootstrap your data file:
    python data/create_sample_data.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def create_sample():
    wb = openpyxl.Workbook()

    # ── Login sheet ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Login"

    headers = ["username", "password", "expected_role", "should_pass"]
    data = [
        ["alice",  "pass_alice",  "Admin", True],
        ["bob",    "pass_bob",    "User",  True],
        ["carol",  "pass_carol",  "User",  True],
        ["hacker", "wrong_pass",  "",      False],
    ]

    # Style header row
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    out = Path("data/test_data.xlsx")
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"  ✅  Created: {out}  ({len(data)} data rows, sheet='Login')")
    print("  Edit this file to add your real test users and expected values.")


if __name__ == "__main__":
    create_sample()
